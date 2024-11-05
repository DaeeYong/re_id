from openpyxl import load_workbook, Workbook
import json
from tqdm import tqdm
import time

# Variables #
#Const
op_xlsx_path = '/Users/ivory/Documents/1/openpose_data/excel/pp151/'
roi_path = '/Users/ivory/Documents/1/pre_roi/pp151/'

'''
task_list = [
    'gait1_front',
    'gait1_rear',
    'gait2_front',
    'gait2_rear',
    'obstacle_high_front',
    'obstacle_high_rear',
    'obstacle_low_front',
    'obstacle_low_rear',
    'tug_front',
    'tug_low',
    'walk_fast_front',
    'walk_fast_rear',
    'walk_preferred_front',
    'walk_preferred_rear',
    'walk_slow_front',
    'walk_slow_rear'
]

'''
task_list = [
    'pp151_omc_gait1_front',
    'pp151_omc_gait2_front',
    'pp151_omc_obstacle_high_front',
    'pp151_omc_obstacle_low_front',
    'pp151_omc_tug_front',
    'pp151_omc_walk_fast_front',
    'pp151_omc_walk_fast_rear',
    'pp151_omc_walk_preferred_front',
    'pp151_omc_walk_preferred_rear',
    'pp151_omc_walk_slow_front',
    'pp151_omc_walk_slow_rear',
    'pp151_omc_walk_turn_front',
    'pp151_omc_walk_turn_rear'
]
#####################

# fuction part #
## private function ##
def _is_point_in_roi(roi:tuple, point: tuple):
    """
    point가 ROI 내부에 존재하는지 판별
    Args:
        roi(tuple): (pt1_x, pt1_y, pt2_x, pt2_y)
        point(tuple): (x, y)
    
    Returns:
        int: point가 ROI내부에 존재하면 1, 그렇지 않으면 0.
    """
    pt1_x, pt1_y, pt2_x, pt2_y = roi
    x, y = point
    
    if pt1_x <= x <= pt2_x and pt1_y <= y <= pt2_y:
        return 1
    else:
        return 0
def _find_idx(data, target):
    """
    Args:
        data: 2차원 배열
        target: 타겟 숫자
    """
    for i, row in enumerate(data):
        if row[0] == target:
            return i
    return -1

## public finction ##
def re_identification2(roi_dict: dict, op_data_wb: Workbook, sheet_name_list: list[str]) -> dict:
    """
    ROI data를 이용해서 OpenPose의 people idx를 재분류하는 함수.

    Args:
        roi_dict (dict): 각 프레임에 대한 bounding box 정보.
        op_data_wb (Workbook): OpenPose 데이터를 포함한 엑셀 워크북.
        sheet_name_list (list[str]): 워크북의 시트 이름 목록.

    Returns:
        dict: ROI별로 재분류된 프레임 데이터.
    """
    reID_dict = {'0': [], '1': []}
    op_dict = {}

    # 각 시트에서 데이터 읽기
    for cur_sheet in sheet_name_list:
        ws = op_data_wb[cur_sheet]
        op_dict[cur_sheet] = [list(row) for row in ws.iter_rows(values_only=True)]

    # ROI별로 데이터 처리
    for roi_key in roi_dict:
        for cur_roi in roi_dict[roi_key]:
            cur_roi_frame_number, pt1_x, pt1_y, pt2_x, pt2_y = cur_roi

            op_candidate = []
            for sheet_name in sheet_name_list:
                op_idx = _find_idx(op_dict[sheet_name], cur_roi_frame_number)
                if op_idx != -1:  # 프레임 번호가 시트에 존재하는 경우에만 추가
                    op_candidate.append(op_dict[sheet_name][op_idx])

            if not op_candidate:  # 후보가 없으면 다음 ROI로 이동
                continue

            # joint 위치가 가장 많이 포함된 후보 찾기
            cnt_list = [0] * len(op_candidate)  # 후보 목록의 조인트 포함 개수 저장
            for idx, each_row in enumerate(op_candidate):
                for i in range(1, 75, 3):  # x, y 좌표만 검사
                    joint_x = each_row[i]
                    joint_y = each_row[i + 1]
                    if joint_x is not None and joint_y is not None:  # None 값 확인
                        cnt_list[idx] += _is_point_in_roi((pt1_x, pt1_y, pt2_x, pt2_y), (joint_x, joint_y))

            if cnt_list and max(cnt_list) > 0:  # 최대 포함 개수가 0보다 클 때만 추가
                max_idx = cnt_list.index(max(cnt_list))
                reID_dict[roi_key].append(op_candidate[max_idx])

    return reID_dict

def re_identification(roi_dict:dict, op_data_wb:Workbook, sheet_name_list:list[str]):
    """
    ROI data를 이용해서 OpenPose의 people idx를 재분류하는 함수.
    arg:
        roi_dict (dict): 각 프레임에 대한 bounding box. idx는 0또는 1
            format -> {
                "{0 || 1}":[
                    [0, x1, y1, x2, y2],
                    [1, x1, y1, x2, y2],
                    ...
                    [t-1, x1, y1, x2, y2]
                ]
            }
        op_data_wb (Workbook):
            data format -> 
                row-1: [0, 25 * 3], 
                row-2: [1, 25 * 3],
                row-3: [2, 25 * 3],
                ...
                row-t: [t-1, 25 * 3]
            
        sheet_name_list (list[str]): ["sheet1", "sheet2", ...]
        
        threshold (int): 분류 기준 임계치
    return:
        reID_dict (dict)
            format -> {
                "0": [
                    [0, 25 * 3],
                    [1, 25 * 3],
                    ..
                    [t-1, 25 * 3]
                ]
                "1": [
                    ...
                ]
            }
    """
    reID_dict = {'0':[], '1':[]}
    op_dict = {}
    '''
    op_dict = {
        sheet_name: [[frmae number, 25*3], [], []... []],
        sheet_name: [[frmae number, 25*3], [], []... []],
        sheet_name: [[frmae number, 25*3], [], []... []],
                }
    '''
    
    roi0_data_list = roi_dict['0']
    roi1_data_list = roi_dict['1']
    
    #op_data_wb를 dict 형태로 변환
    for cur_sheet in sheet_name_list:
        ws = op_data_wb[cur_sheet]
        op_dict[cur_sheet] = []
        
        for row in ws.iter_rows(values_only=True):
            op_dict[cur_sheet].append(list(row))
    ############################
    
    #roi idx 0번
    for cur_roi in roi0_data_list:
        cur_roi_frame_number, pt1_x, pt1_y, pt2_x, pt2_y = cur_roi
        
        op_candidate = []
        for sheet_name in sheet_name_list:
            op_idx = _find_idx(op_dict[sheet_name], cur_roi_frame_number)
            if op_idx != -1:
                op_candidate.append(op_dict[sheet_name][op_idx])
        
        #현재 op_candidate에 후보 목록들이 들어가 있는 상태 e.g. [[sheet1],[sheet2],[sheet3],[sheet4]]
        #지금부터 joint pose가 몇개 들어가 있는지 세야함.
        cnt_list = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #각 sheet의 row joint 들어가는 개수
        for idx, each_row in enumerate(op_candidate):
            for i in range(1, 75, 3):
                joint_x = each_row[i]
                joint_y = each_row[i+1]
                
                cnt_list[idx] +=_is_point_in_roi((pt1_x, pt1_y, pt2_x, pt2_y), (joint_x, joint_y))
        
        max_idx = cnt_list.index(max(cnt_list))
        reID_dict['0'].append(op_candidate[max_idx])
        
    #roi idx 1번
    for cur_roi in roi1_data_list:
        cur_roi_frame_number, pt1_x, pt1_y, pt2_x, pt2_y = cur_roi
    
        op_candidate = []
        for sheet_name in sheet_name_list:
            op_idx = _find_idx(op_dict[sheet_name], cur_roi_frame_number)
            if op_idx != -1:
                op_candidate.append(op_dict[sheet_name][op_idx])
        
        #현재 op_candidate에 후보 목록들이 들어가 있는 상태 e.g. [[sheet1],[sheet2],[sheet3],[sheet4]]
        #지금부터 joint pose가 몇개 들어가 있는지 세야함.
        cnt_list = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #각 sheet의 row joint 들어가는 개수
        for idx, each_row in enumerate(op_candidate):
            for i in range(1, 75, 3):
                joint_x = each_row[i]
                joint_y = each_row[i+1]
                
                cnt_list[idx] +=_is_point_in_roi((pt1_x, pt1_y, pt2_x, pt2_y), (joint_x, joint_y))
        
        max_idx = cnt_list.index(max(cnt_list))
        reID_dict['1'].append(op_candidate[max_idx])
            
    return reID_dict
    
#####################

# logic #
# 기존 엑셀 파일 및 ROI 데이터 불러오기
for TASK_IDX, _ in enumerate(task_list):
    print(_)
    joint_data_path = op_xlsx_path + task_list[TASK_IDX] + '.xlsx'
    roi_json_path = roi_path +task_list[TASK_IDX] + '.json'
    
    xlsx_sheet_name_list = []
    with open(roi_json_path, 'r') as file:
        roi_dict = json.load(file)

    op_wb = load_workbook(joint_data_path)
    xlsx_sheet_name_list = op_wb.sheetnames

    # ROI 데이터를 기반으로 재분류된 데이터 생성
    reID_dict = re_identification(roi_dict, op_wb, xlsx_sheet_name_list)

    # reID_dict 내용을 새로운 시트(reid0, reid1)에 저장
    reid_sheet_name_list = ['reid0', 'reid1']
    for idx, re_idx_name in enumerate(reid_sheet_name_list):
        # 새 시트 생성 (기존 시트 존재 여부는 확인하지 않음)
        sheet = op_wb.create_sheet(title=re_idx_name)
        
        # 시트에 데이터 추가
        for row in reID_dict[str(idx)]:  # str(idx)로 키에 접근
            sheet.append(row)

    # 변경된 내용을 파일에 저장
    op_wb.save(joint_data_path)