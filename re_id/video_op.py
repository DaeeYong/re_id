import cv2
from openpyxl import load_workbook
import copy

####Variables####
#Const
op_xlsx_path = '/Users/ivory/Documents/1/openpose_data/excel/pp151/'
op_joint_path = '/Users/ivory/Documents/1/openpose_data/2d_joint/'
pp11_video_path = '/Users/ivory/Documents/1/pp151/'
'''
task_list = [
    'gait1_front',
    'gait1_rear',
    'gait2_front',
    'gait2_rear',
    'obstacle_high_front',
    'obstacle_high_rear',
    'obstacle_low_front',
    'obstacle_low_rear',
    'tug_front',
    'tug_low',
    'walk_fast_front',
    'walk_fast_rear',
    'walk_preferred_front',
    'walk_preferred_rear',
    'walk_slow_front',
    'walk_slow_rear'
]

'''
task_list = [
    'pp151_omc_gait1_front',
    'pp151_omc_gait2_front',
    'pp151_omc_obstacle_high_front',
    'pp151_omc_obstacle_low_front',
    'pp151_omc_tug_front',
    'pp151_omc_walk_fast_front',
    'pp151_omc_walk_fast_rear',
    'pp151_omc_walk_preferred_front',
    'pp151_omc_walk_preferred_rear',
    'pp151_omc_walk_slow_front',
    'pp151_omc_walk_slow_rear',
    'pp151_omc_walk_turn_front',
    'pp151_omc_walk_turn_rear'
]

ACTIVE_SHEET_NAME = ['OpIdx0','OpIdx1', 'OpIdx2', 'OpIdx3','reid0', 'reid1']
#controll variables
TASK_IDX = 0
selected = 4

joint_data_path = op_xlsx_path + task_list[TASK_IDX] + '.xlsx'
video_path = pp11_video_path + task_list[TASK_IDX] + '.mp4'
#####################

#function part#
def play_video_with_joint(video_path:str, joint_data_list, video_name:str = 'video', play_mul:float = 1):
    """
    openpose의 2d Joint data를 video에 렌더링
    args:
        video_path (str): video path
        joint_data_list(list[[]]): 2차원 배열 [[frame_num, #25x2]]
    return:
    
    """
    #Const variables
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)
    op_pointer = 0
    
    if not cap.isOpened():
        print("Error: Could not open video.")
        exit()
        
    while True:
        ret, frame = cap.read()
        if not ret:
            break
    
        #algorithm space
        ##variabels
        cur_frame_num = int(cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
        op_frame_num = joint_data_list[op_pointer][0]
        cur_joint_data = joint_data_list[op_pointer][1:]
        
        ##logic
        if(cur_frame_num == op_frame_num):
            print(f'{cur_frame_num}, {op_frame_num}')
            
            for i in range(0, len(cur_joint_data), 2):
                x = int(cur_joint_data[i])
                y = int(cur_joint_data[i+1])
                
                center_coord = (x, y)
                cv2.circle(frame, center_coord, radius=6, color=(0, 0, 255), thickness=-1)
            
            op_pointer += 1
            
        #global text part
        text = f'Frame: {cur_frame_num}'
        cv2.putText(frame, text, (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0) , 3)
        #############
        
        #play video#
        cv2.imshow(video_name, frame)
        if cv2.waitKey(int(25 / play_mul)) & 0xFF == ord('q'):
            break
    
    cap.release()
    cv2.destroyAllWindows()

####################

#logic part#
wb = load_workbook(joint_data_path)
ws = wb[ACTIVE_SHEET_NAME[selected]]

xlsx_data_list = []
for row in ws.iter_rows(values_only=True):
    xlsx_data_list.append(list(row))

all_frame_joint_list = []
for each_row in xlsx_data_list:
    now_frame_data = []
    row_len = len(each_row)
    #frame number
    now_frame_data.append((each_row[0]))
    #each row에 대한 처리
    for i in range(1, row_len, 3):
        x = each_row[i]
        y = each_row[i+1]
        
        now_frame_data.append(x)
        now_frame_data.append(y)
        
    all_frame_joint_list.append(copy.deepcopy(now_frame_data))

play_video_with_joint(video_path, all_frame_joint_list)