import cv2
import os
import json
import openpyxl as xl
import pandas as pd
from sklearn.preprocessing import MinMaxScaler
import numpy as np
POSE_2D_JOINT_NAME_LIST = [
        'Nose_x', 'Nose_y', 
        'Neck_x', 'Neck_y', 
        'RShoulder_x', 'RShoulder_y', 
        'RElbow_x', 'RElbow_y', 
        'RWrist_x', 'RWrist_y', 
        'LShoulder_x', 'LShoulder_y', 
        'LElbow_x', 'LElbow_y', 
        'LWrist_x', 'LWrist_y', 
        'MidHip_x', 'MidHip_y', 
        'RHip_x', 'RHip_y',
        'RKnee_x', 'RKnee_y',
        'RAnkle_x', 'RAnkle_y', 
        'LHip_x', 'LHip_y',
        'LKnee_x', 'LKnee_y', 
        'LAnkle_x', 'LAnkle_y',
        'REye_x', 'REye_y',
        'LEye_x', 'LEye_y', 
        'REar_x', 'REar_y', 
        'LEar_x', 'LEar_y', 
        'LBigToe_x', 'LBigToe_y', 
        'LSmallToe_x', 'LSmallToe_y', 
        'LHeel_x', 'LHeel_y',
        'RBigToe_x', 'RBigToe_y',
        'RSmallToe_x', 'RSmallToe_y', 
        'RHeel_x', 'RHeel_y'
]

def add_number(a:int, b:int) -> int:
    return a+b

def remove_confidence_from_keypoints_2d(p_keypoints):
    return [value for index, value in enumerate(p_keypoints, start=1) if index % 3!= 0]

def nomalize_xlsx_and_save(frame_data):
    pass


def nomalize_data(frame_list):
    scalar = MinMaxScaler()
    nomalized_frame_list = scalar.fit_transform(frame_list).tolist()
    
    return nomalized_frame_list

def framelist2excel(frame_list, save_path):
    wb = xl.Workbook()
    sheet = wb.active

    #sheet.append(feature_list)
    for each_frame in frame_list:
        sheet.append(each_frame)
    
    wb.save(save_path)
    print(f'[save done] path : {save_path}\n')
    
def mark_pos(img, x, y, radius = 5, color = (0, 0, 255), thickness = 2):
    _x = round(x)
    _y = round(y)
    cv2.circle(img, (_x, _y), radius, color, thickness)

def get_total_frame(video_path):
    cap = cv2.VideoCapture(video_path)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    cap.release()
    
    return total_frames

def get_jsons_list(folder_path):
    file_list = [f for f in sorted(os.listdir(folder_path)) if f.endswith('.json')]
    return file_list

def get_all_frame_data_from_jsons_list(jsons_list:list, root_dir:str) -> list :
    #3차원 배열
    all_frame_data = []

    for now_json in jsons_list:
        with open(root_dir + now_json, 'r') as file:
            data = json.load(file)
        
        people = data['people']
        len_people = len(people)

        now_frame = []
        for idx in range(0, len_people):
            p = people[idx]
            p_keypoints_2d = p['pose_keypoints_2d']

            p_keypoints_2d = remove_confidence_from_keypoints_2d(p_keypoints_2d)
            now_frame.append(p_keypoints_2d)
        
        all_frame_data.append(now_frame)

    return all_frame_data

#this function is for only openpose json body25 model
def jsons2excel(json_dir_path, feature_list, key_index, file_name, save_path, json_key = "people"):
    #create xlsx file
    workbook = xl.Workbook()
    sheet = workbook.active
    
    #feature_name
    #sheet.append(feature_list)

    #fetch json file list
    json_list = get_jsons_list(json_dir_path)
    json_list.sort()
    #read json file sequentially
    for name in json_list:
        #read json file
        with open(json_dir_path + name, 'r') as file:
            data = json.load(file)

        # json file에 'people'의 value가 비어있는 경우
        if data['people'] == []:
            empty_list = [0] * 50
            sheet.append(empty_list)
            continue
        
        people = data[json_key]
        # people 요소 길이와 key idx 비교
        if len(people) -1 < key_index:
            empty_list = [0] * 50
            sheet.append(empty_list)
            continue

        p = people[key_index]
        p_keypoints = p["pose_keypoints_2d"]
        
        # except confidence value
        p_coord = [value for index, value in enumerate(p_keypoints, start=1) if index % 3!= 0]

        #save json data at each row
        sheet.append(p_coord)

    workbook.save(save_path + file_name)

'''
data_path : data type is only axcel available.
feature_list : feature_list of data. the type is list(str).
video_path : raw video path
save_video_path : processed video path where you want to save 
'''
#speed unit : ms
def mark_position_at_video(data_path, video_path, video_name, speed = 100):
    cap = cv2.VideoCapture(video_path)
    total_frame = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # read pos data
    wb = xl.load_workbook(data_path, data_only=True)
    #pos_sheet = wb.active
    pos_sheet = wb['frame']
    frame_pos_data_list = []
    
    for row in pos_sheet.iter_rows(values_only=True):
        frame_pos_data_list.append(list(row))

    num_label = len(frame_pos_data_list[0])
    
    #open video
    if not cap.isOpened():
        print("Error: Could not open video")
        return -1
    
    print(f"[read success ]{video_path}")

    while cap.isOpened():
        for frame_num in range(0, total_frame):
            ret, frame = cap.read()

            if not ret:
                break
            # processing area #
            for col in range(0, num_label, 2):
                x = frame_pos_data_list[frame_num + 1][col]
                y = frame_pos_data_list[frame_num + 1][col + 1]

                mark_pos(frame, x, y)
            ####################
            print(f"now frame : {frame_num}\n")
            cv2.imshow(video_name, frame)
            cv2.waitKey(speed)
        
        if cv2.waitKey(30) & 0xFF == 27:
            break

    cap.release()
    cv2.destroyAllWindows()

def xlsx2data(data_path, sheet_name='Sheet', start_row = 1):
    wb = xl.load_workbook(data_path, data_only=True)
    sheet = wb[sheet_name]
    frame_data_list = []

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        frame_data_list.append(list(row))
    
    return frame_data_list


def get_poskeypoints2d_from_json(file_path:str, people_idx:int)-> list:
    with open(file_path, 'r') as file:
        data = json.load(file)
        people = data['people']
        
        p = people[people_idx]
        p_keypoints = p["pose_keypoints_2d"]

        return [value for index, value in enumerate(p_keypoints, start=1) if index % 3!= 0]

def play_video(video_path, video_name):
    cap = cv2.VideoCapture(video_path)

    if not cap.isOpened():
        print(f'[Error: could not open video]:{video_path}')
        return -1
    
    print(f"[read success]{video_path}")
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break
    
        #현재 프레임 번호 출력
        current_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES))
        print(f'current_frame : {current_frame}')

        cv2.imshow(video_name, frame)        
        if cv2.waitKey(25) & 0xFF == 27:  # esc 키의 ASCII 코드는 27
            break

    cap.release()
    cv2.destroyAllWindows()

def show_specific_video_frame(video_path, video_name : str, frame_num = 1):
    cap = cv2.VideoCapture(video_path)
    frame_number = frame_num
    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
    ret, frame = cap.read()
    # 비디오 파일 닫기
    cap.release()

    # 프레임 표시
    cv2.imshow(video_name, frame)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

def get_specific_video_frame(video_path:str, frame_num = 0) -> np.ndarray:
    cap = cv2.VideoCapture(video_path)
    frame_number = frame_num
    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
    ret, frame = cap.read()
    # 비디오 파일 닫기
    cap.release()
    
    if ret == False:
        return False
    
    return frame

# 최신판
def mark_pos_on_video(video_path:str, frame_data_list : list, video_name, start_frame_idx = 0, speed = 1):
    flag = False
    cap = cv2.VideoCapture(video_path)

    
    frame_data_max_idx = len(frame_data_list) - 1
    
    if not cap.isOpened():
        print(f'[Error: could not open video]:{video_path}')
        return -1

    for i in range(0, start_frame_idx + 1):
        ret, frame = cap.read()

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break
        
        #현재 프레임 번호
        current_frame_idx = int(cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1

        if current_frame_idx < frame_data_max_idx + 1:
            #하나의 프레임에 해당하는 위치 그리기
            for i in range(0, 50, 2):
                x = frame_data_list[current_frame_idx][i]
                y = frame_data_list[current_frame_idx][i + 1]

                mark_pos(frame, x, y)
        
        ################
        print(f'now frame idx: {current_frame_idx}')
        cv2.imshow(video_name, frame)
        cv2.waitKey(speed)
        if cv2.waitKey(30) & 0xFF == 27:
            cap.release()
            cv2.destroyAllWindows()
            exit()
        

    cap.release()
    cv2.destroyAllWindows()

# label을 rendering 하는 함수
# gt = [ ltoe, lheel, rtoe, rheel ], [ 19, 21, 22, 24 ]
def render_result_on_video(video_path:str, joint_gt_pair_list : list, video_name:str, speed = 1):
    label_number = [19, 21, 22, 24]
    cap = cv2.VideoCapture(video_path)
    frame_data_max_idx = len(joint_gt_pair_list) - 1

    if not cap.isOpened():
        print(f'[Error: could not open video]:{video_path}')
        return -1

    print(f"[read success]{video_path}")
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break
        
        #현재 프레임 번호
        current_frame_idx = int(cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
        
        if current_frame_idx < frame_data_max_idx + 1:
            #모든 joint pos plotting
            for i in range(0, 50, 2):
                x = joint_gt_pair_list[current_frame_idx][0][i]
                y = joint_gt_pair_list[current_frame_idx][0][i + 1]

                mark_pos(frame, x, y, color=(0, 0, 255))

            #label parsing
            ltoe_label = joint_gt_pair_list[current_frame_idx][1][0]
            lheel_label = joint_gt_pair_list[current_frame_idx][1][1]
            rtoe_label = joint_gt_pair_list[current_frame_idx][1][2]
            rheel_label = joint_gt_pair_list[current_frame_idx][1][3]

            ###label rendering###
            
            if ltoe_label == 1:
                x = joint_gt_pair_list[current_frame_idx][0][label_number[0] * 2]
                y = joint_gt_pair_list[current_frame_idx][0][label_number[0] * 2 + 1]
                mark_pos(frame, x, y, color=(0, 255, 0))
            
            if lheel_label == 1:
                x = joint_gt_pair_list[current_frame_idx][0][label_number[1] * 2]
                y = joint_gt_pair_list[current_frame_idx][0][label_number[1] * 2 + 1]
                mark_pos(frame, x, y, color=(0, 255, 0))

            if rtoe_label == 1:
                x = joint_gt_pair_list[current_frame_idx][0][label_number[2] * 2]
                y = joint_gt_pair_list[current_frame_idx][0][label_number[2] * 2 + 1]
                mark_pos(frame, x, y, color=(0, 255, 0))

            if rheel_label == 1:
                x = joint_gt_pair_list[current_frame_idx][0][label_number[3] * 2]
                y = joint_gt_pair_list[current_frame_idx][0][label_number[3] * 2 + 1]
                mark_pos(frame, x, y, color=(0, 255, 0))

        ################
        print(f'now frame idx: {current_frame_idx}')
        cv2.imshow(video_name, frame)
        cv2.waitKey(speed)
        if cv2.waitKey(30) & 0xFF == 27:
            break

    cap.release()
    cv2.destroyAllWindows()

'''
데이터 반환 형식 : 3차원 배열
[
    [ [ each_frame_data ],[ label ] ],
    [ [ each_frame_data ],[ label ] ],
    [ [ each_frame_data ],[ label ] ],
                ...
    [ [ each_frame_data ],[ label ] ],
    [ [ each_frame_data ],[ label ] ]
]
'''
def make_dataAndGtPair(all_frame_data : list, label : list, startidx = 1) -> list:
    frame_len = len(all_frame_data)
    label_len = len(label)

    data_gt_pair_list = []
    
    shorter_len = 0
    if frame_len < label_len : shorter_len = frame_len
    else : shorter_len = label_len

    for idx in range(startidx, shorter_len):
        tmp = []
        tmp.append(all_frame_data[idx])
        tmp.append(label[idx])
        data_gt_pair_list.append(tmp)
    
    return data_gt_pair_list

'''
selected_joint_number_list = [1, 2, 3, 10, 24]
'''
def get_selected_joint_pos_frame_list(joint_frame_data : list, selected_joint_number_list : list):
    selected_joint_frame_data_list = []

    frame_len = len(joint_frame_data)
    for idx in range(0, frame_len):
        each_frame = []
        for joint_idx in selected_joint_number_list:
            x = joint_frame_data[idx][joint_idx * 2]
            y = joint_frame_data[idx][joint_idx * 2 + 1]
            each_frame.append(x)
            each_frame.append(y)

        selected_joint_frame_data_list.append(each_frame)

    return selected_joint_frame_data_list

def get_videoName_list(folder_path, video_extensions=['.mp4', '.avi', '.mkv', '.mov', '.flv', '.wmv']):
    file_list = os.listdir(folder_path)
    video_files = [file for file in file_list if os.path.splitext(file)[1].lower() in video_extensions]

    return video_files

def get_video_durationSec(video_path):
    cap = cv2.VideoCapture(video_path)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps = cap.get(cv2.CAP_PROP_FPS)

    duration = total_frames / fps
    
    cap.release()

    return round(duration)

if __name__=="__main__":
    print(os.__version__)
